#----------------------------------------------------------
#Streamflow Catalogue Master File
#Kaiser, Blasch and Schmitz
#2023, Boise State University, Department of Geology
#----------------------------------------------------------

#----------------------------------------------------------
#import section (external Python modules adding additional functionality)
#openpyxl   - handles excel,csv documents
#matplotlib - data visualizer
#numpy      - expands Python base math functionality
#datetime   - Python date handler
#----------------------------------------------------------

from datetime import datetime
import openpyxl as px
import matplotlib.pyplot as plt
import numpy as np
import matplotlib as mpl

#----------------------------------------------------------
#importing streamflow catalogue
#----------------------------------------------------------
print('loading catalog...')
cat_dir = 'C:/Users/sjsch/Desktop/Kendra/Streamflow_Catalog.xlsx' #be sure to set location
workbook = px.open(cat_dir)
wb = workbook.active

#----------------------------------------------------------
#Determine length of Streamflow Catalog as it constantly updates
#----------------------------------------------------------

cat_total = 0
maxrow = 1
for val in wb.iter_rows(min_col=1,max_col=1):
    while wb.cell(maxrow,1).value is not None:
        cat_total += 1
        maxrow += 1
maxrow = maxrow - 1
print(f"Streamflow catalog is {maxrow} lines")
#----------------------------------------------------------
#establishing histogram figure 1 sorting parameters
#----------------------------------------------------------
years_start = []
years_end = []
unknown = []
counter = 1
try:
    for row in wb.iter_rows(min_col=15,max_col=15,min_row=2,max_row=maxrow): #edit min,max and figure out on the fly calc
        for cell in row:
            counter += 1
            date_format = ('%Y-%m-%d %H:%M:%S')
            date_format1 = ('%H:%M:%S')
            var = str(cell.value).lower()
            if str(wb.cell(counter,16).value) == 'current':
                cell.value = datetime.now()
                pass
            elif str(wb.cell(counter,1).value).lower() == 'none':
                continue
            elif str(wb.cell(counter,21).value).lower() == 'continuous':
                if var == 'none':
                    if str(wb.cell(counter,16).value) == 'None':
                        continue
                    else:
                        unknown.append(var)
                        continue
                elif var == 'unknown':
                    unknown.append(var)
                    continue
                elif var == '2/1/1857':
                    years_start.append('1857')
                    years_end.append('2022')
                elif str(wb.cell(counter,15).value) == '00:00:00':
                    continue
                else:
                    if str(wb.cell(counter,16).value) == 'None' or str(wb.cell(counter,16).value).lower() == 'unknown':
                        unknown.append(var)
                        continue
                    date_s = datetime.strptime(var,date_format)
                    years_start.append(date_s.year)
                    date_e = datetime.strptime(str(wb.cell(counter,16).value),date_format)
                    years_end.append(date_e.year)
                    continue
            else:
                pass
except:
    print(counter)
    raise
#----------------------------------------------------------
#determining figure data range
#----------------------------------------------------------
totals = []
counter = 0
try:
    for i in range(0,maxrow):
        var = int(years_end[counter]) - int(years_start[counter])
        totals.append(var)
        counter += 1
except:
    pass

#----------------------------------------------------------
#setting figure arguments, parameters for histogram
#----------------------------------------------------------
dist1 = totals
dist2 = range(0,126,10)
fig, axs = plt.subplots(tight_layout=True)
axs.hist(dist1,bins=125,range=(0,125))
axs.set_xticks(dist2)
axs.set_ylabel('quantity')
axs.set_xlabel('years of operation')
axs.set_xlim([0,125])
plt.show()

#----------------------------------------------------------
#establishing active/inactive figure 2 parameters
#----------------------------------------------------------

OR_inact,OR_act = [],[]
ID_inact,ID_act = [],[]
WA_inact,WA_act = [],[]
ID_unk,OR_unk,WA_unk = [],[],[]

#function to sort by state and active designation
def ActiveCounter():
    counter = 1
    for row in wb.iter_rows(min_col=21,max_col=21,min_row=2,max_row=maxrow):
        for cell in row:
            var = str(cell.value).lower()
            counter += 1
            if var == 'continuous' or var == 'seasonal': #subsets by 'continuous' designation
                if str(wb.cell(counter,6).value) == 'Idaho':
                    if str(wb.cell(counter,17).value) == 'active':
                        ID_act.append(var)
                    elif str(wb.cell(counter,17).value) == 'inactive':
                        ID_inact.append(var)
                    else:
                        ID_unk.append(var)
                elif str(wb.cell(counter,6).value) == 'Oregon':
                    if str(wb.cell(counter,17).value) == 'active':
                        OR_act.append(var)
                    elif str(wb.cell(counter,17).value) == 'inactive':
                        OR_inact.append(var)
                    else:
                        OR_unk.append(var)
                elif str(wb.cell(counter,6).value) == 'Washington':
                    if str(wb.cell(counter,17).value) == 'active':
                        WA_act.append(var)
                    elif str(wb.cell(counter,17).value) == 'inactive':
                        WA_inact.append(var)
                    else:
                        WA_unk.append(var)
ActiveCounter()                    
print(
    f"Idaho: {len(ID_inact)} - inactive, {len(ID_act)} - active\n"\
    f"Oregon: {len(OR_inact)} - inactive, {len(OR_act)} - active\n"\
    f"Washington: {len(WA_inact)} - inactive, {len(WA_act)} - active")

#designating graphic x,y parameters
labels = ['Washington','Oregon','Idaho']
active = [len(WA_act),len(OR_act),len(ID_act)]
inactive = [len(WA_inact),len(OR_inact),len(ID_inact)]
unknown = [len(WA_unk),len(OR_unk),len(ID_unk)]
x = np.arange(len(active))
width = 0.3

#chart graphic
fig, ax = plt.subplots()
rects1 = ax.bar(x - 0.15,active, width, label='active',color='deepskyblue')
rects2 = ax.bar(x + 0.15, inactive, width, label='inactive',color='orange')
rects3 = ax.bar(x + 0.45, unknown, width, label='unknown',color='black')
ax.yaxis.set_major_formatter(mpl.ticker.StrMethodFormatter('{x:,.0f}')) #formats into #,### integers
ax.set_ylabel('Number of Continuous Gages')
ax.set_xticks(x, labels)

#defines integer labels for each bar, pulls directly from graph data
ax.bar_label((rects1), padding=1)
ax.bar_label((rects2), labels=[f'{x:,.0f}' for x in rects2.datavalues], padding=1)
ax.bar_label((rects3), labels=[f'{x:,.0f}' for x in rects3.datavalues], padding=1)
ax.legend(fancybox=True)

fig.tight_layout()
plt.show()

OR_cont,OR_misc,OR_unk = [],[],[]
WA_cont,WA_misc,WA_unk = [],[],[]
ID_cont,ID_misc,ID_unk = [],[],[]
counter = 1

#----------------------------------------------------------
#establishing figure 3 parameters
#----------------------------------------------------------
for row in wb.iter_rows(min_row=2,max_row=maxrow,min_col=21,max_col=21):
    for cell in row:
        counter += 1
        if str(wb.cell(counter,2).value) == 'USGS':
            continue
        var = str(cell.value).lower()
        if var == '' or var == 'none' or var == 'unknown':
            if str(wb.cell(counter,6).value) == 'Idaho':
                ID_unk.append(var)
            elif str(wb.cell(counter,6).value) == 'Oregon':
                OR_unk.append(var)
            elif str(wb.cell(counter,6).value) == 'Washington':
                WA_unk.append(var)
        elif var == 'continuous':
            if str(wb.cell(counter,6).value) == 'Idaho':
                ID_cont.append(var)
            elif str(wb.cell(counter,6).value) == 'Oregon':
                OR_cont.append(var)
            elif str(wb.cell(counter,6).value) == 'Washington':
                WA_cont.append(var)
        elif var == 'seasonal' or var == 'synoptic':
            if str(wb.cell(counter,6).value) == 'Idaho':
                ID_misc.append(var)
            elif str(wb.cell(counter,6).value) == 'Oregon':
                OR_misc.append(var)
            elif str(wb.cell(counter,6).value) == 'Washington':
                WA_misc.append(var)
                
        else:
            pass

#----------------------------------------------------------
#calculating category totals
#----------------------------------------------------------
ID_tot = len(ID_cont) + len(ID_misc) + len(ID_unk)
OR_tot = len(OR_cont) + len(OR_misc) + len(OR_unk)
WA_tot = len(WA_cont) + len(WA_misc) + len(WA_unk)

unk_tot = (len(ID_unk),len(OR_unk),len(WA_unk))
misc_tot = (len(ID_misc),len(OR_misc),len(WA_misc))
cont_tot = (len(ID_cont),len(OR_cont),len(WA_cont))

categories = ['Unknown','Miscellaneous','Continuous(%)']
gages_no = ['Idaho : %s' % ID_tot,'Oregon : %s' % OR_tot,'Washington : %s' % WA_tot]
results = { 
    str(ID_tot):[round((len(ID_unk)/ID_tot)*100,1),round((len(ID_misc)/ID_tot)*100,1),round((len(ID_cont)/ID_tot)*100,1)],
    str(OR_tot):[round((len(OR_unk)/OR_tot)*100,1),round((len(OR_misc)/OR_tot)*100,1),round((len(OR_cont)/OR_tot)*100,1)],
    str(WA_tot):[round((len(WA_unk)/WA_tot)*100,1),round((len(WA_misc)/WA_tot)*100,1),round((len(WA_cont)/WA_tot)*100,1)]
    }
results2 = {
    ID_tot:[ID_cont,ID_misc,ID_unk],
    OR_tot:[OR_cont,OR_misc,OR_unk],
    WA_tot:[WA_cont,WA_misc,WA_unk]
    }
labeltotals = (("{:,}".format(ID_tot),"{:,}".format(OR_tot),"{:,}".format(WA_tot))) #format integers into #,###

#----------------------------------------------------------
#graphics conditions
#----------------------------------------------------------
def PlotDat(results, categories):
    ylabels = np.arange(len(categories))
    labels = list(results.keys())
    data = np.array(list(results.values()))
    data_cum = data.cumsum(axis=1)
    category_colors = plt.colormaps['Wistia'](
        np.linspace(0.15, 0.85, data.shape[1]))
    fig, ax = plt.subplots(figsize=(9.2, 5))
    ax.xaxis.set_visible(False)
    ax.set_xlim(0, np.sum(data, axis=1).max())
    for i, (colname, color) in enumerate(zip(categories, category_colors)): #bar generator for 3x3 distribution
        widths = data[:,i]
        starts = data_cum[:,i] - widths
        rects = ax.barh(labels,widths,left=starts,height=0.5,label=colname,color=color)
        ax.yaxis.set_major_formatter(mpl.ticker.StrMethodFormatter('{x:,.0f}'))
        #ax.xaxis.set_major_formatter(mpl.ticker.StrMethodFormatter('{x:,.0f}'))#formats into #,### integers
        r,g,b,_ = color
        text_color = 'black'
        ax.bar_label(rects,label_type='center',wrap=False,color=text_color,padding=5.5)
    ax.legend(ncol=len(categories), bbox_to_anchor=(0.25,0),
              loc='upper center', fontsize='small')
    ax.legend(ncol=len(categories), bbox_to_anchor=(0.75,0),
              loc='upper center', fontsize='small')
    ax.set_yticks(ylabels)
    ax.set_xticklabels(labeltotals)
    ax.bar_label(rects, gages_no, padding=0.5)
    ax.set_ylabel('quantity')
    return fig, ax

cont1 = [round((len(ID_cont)/ID_tot)*100,0),round((len(OR_cont)/OR_tot)*100,0),round((len(WA_cont)/WA_tot)*100,0)]
misc1 = [round((len(ID_misc)/ID_tot)*100,0),round((len(OR_misc)/OR_tot)*100,0),round((len(WA_misc)/WA_tot)*100,0)]
unk1 = [round((len(ID_unk)/ID_tot)*100,0),round((len(OR_unk)/OR_tot)*100,0),round((len(WA_unk)/WA_tot)*100,0)]

fig, ax = plt.subplots()
width = 0.45
a = ax.bar(gages_no, cont1, width, label='Continuous',color='deepskyblue')
b = ax.bar(gages_no, misc1, width, label='Discrete',bottom=cont1,color='darkorange')
#c = ax.bar(gages_no, unk1, width, label='Unknown',bottom=0,color='black')
ax.bar_label((a), label_type='center', color='black',padding=3)
ax.bar_label((b), label_type='center', color='black',padding=3)
#ax.bar_label((c), label_type='center', color='black',padding=3)
ax.set_ylabel('Percent of Dataset (%)')
ax.legend(ncol=len(categories),bbox_to_anchor=(0, 1),loc='lower left', fontsize='small')
#ax.xaxis.set_major_formatter(mpl.ticker.StrMethodFormatter('{x:,.0f}'))
#txt=("Idaho totals: ",ID_tot," Oregon totals: ",OR_tot," Washington totals: ",WA_tot)
#ax.text(0.5, 0.01, txt, wrap=True, horizontalalignment='center', fontsize=12)
#ax.set_xticklabels(labeltotals)
plt.show()



