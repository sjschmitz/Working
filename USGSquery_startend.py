##########
#Query start/end dates for USGS gages from nwis server and write to streamflow catalog
#last update: 4.27.2023
#
#processes:
#   import Excel file, extract USGS gage id
#   query site data from nwis
#   parse start/end dates
#   write to Excel file
#   FACT CHECK!!
##########



import re
import requests
from dataretrieval import nwis
import pandas as pd
import openpyxl as px

# Define file path, on local
file_path = 'C:/Users/sjsch/Desktop/Kendra/Kyle_USGS.xlsx'
print('Import success')

workbook = px.open(file_path)
wb = workbook.active
counter = 0
USGS_gage_id = []
USGS_dict = {}

#function to extract dates in YYYY-MM-DD format out of cell
def find_dates(string):
    pattern = r"\d{4}-\d{2}-\d{2}"
    dates = re.findall(pattern, string)
    
    if dates:
        first_date = dates[0]
        last_date = dates[-1]
        return first_date, last_date
    else:
        return None, None
    
#retrieve USGS site id's from excel file
for row in wb.iter_rows(min_col=1,max_col=1):
    for cell in row:
        var = str(cell.value).lower()
        counter +=1
        if var == 'usgs':
            USGS_gage_id.append(wb.cell(counter,2).value)
        else:
            continue

#query NWIS server with imported site id
USGS_dat = []
for val in USGS_gage_id:
    site_id = str(val)
    data = nwis.get_dv(sites=site_id,start='1900-01-01') #generic start date
    USGS_dat.append(data)
    dic = {val : data} #dictionary with site ID as key and array as value
    USGS_dict.update(dic)

#iterate through returned USGS arrays and use find_dates() to pull out first
#and last recorded data and then write it to excel file
counter = 0
for row in wb.iter_rows(min_col=2,max_col=2):
    for cell in row:
        counter +=1
        var = cell.value
        if var in list(USGS_dict.keys()):
            arr = str(USGS_dict.get(var))
            arr = find_dates(arr)
            wb.cell(counter,9).value = arr[0] #start date
            wb.cell(counter,10).value = arr[1] #end date
            print('Gage #%s started %s and ended %s' % (var, arr[0], arr[1]))
        else:
            print('skip %s' % var)
            continue

workbook.save('USGS_Updated.xlsx') #updated excel file with start and end dates
