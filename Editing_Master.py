import openpyxl as px             #handles catalog as dataframe
import matplotlib.pyplot as plt   #visualizes data
import numpy as np                #mathematical library with built-in calculation tools
import matplotlib as mpl

w_dir = 'C:/Users/sjsch/Desktop/Kendra/burpdates.xlsx' #streamflow catalog pathway
workbook = px.open(w_dir)
wb = workbook.active

#------------------------------------------------------
#checking for streamflow catalog length
#------------------------------------------------------
cat_total = 0
maxrow = 1
for val in wb.iter_rows(min_col=1,max_col=1):
    while wb.cell(maxrow,1).value is not None:
        cat_total += 1
        maxrow += 1
maxrow = maxrow - 1
print(f"Streamflow catalog is {maxrow} lines")
#------------------------------------------------------
#replacing dates with YYYY with 1/1/YYYY for consistant formatting
#------------------------------------------------------
counter = 0
for rows in wb.iter_rows(min_row=1,max_row=12671,min_col=1,max_col=1): #gage start dates
    for cell in rows:
        counter += 1
        date = str(cell.value)
        cell.value = ('1/1/'+date)
    
counter = 0
for rows in wb.iter_rows(min_row=1,max_row=12671,min_col=2,max_col=2): #gage end dates
    for cell in rows:
        counter += 1
        date = str(cell.value)
        cell.value = ('1/1/'+date)

workbook.save('C:/Users/sjsch/Desktop/Kendra/brpdates.xlsx')

#------------------------------------------------------
#checking for empty latitude/longitude fields
#------------------------------------------------------
counter = 1
checks = []
for rows in wb.iter_rows(min_row=2,max_row=maxrow,min_col=11,max_col=11):
    for cell in rows:
        counter += 1
        if str(cell.value).lower() == 'none':
            checks.append(counter)
        else:
            pass

#------------------------------------------------------
#checking for empty date fields/erroneous default values
#------------------------------------------------------
years_start = []
years_end = []
unknown = []
counter = 1
try:
    for row in wb.iter_rows(min_col=6,max_col=6,min_row=2,max_row=maxrow):
        for cell in row:
            counter += 1
            date_format = ('%Y-%m-%d %H:%M:%S')
            date_format1 = ('%H:%M:%S')
            var = str(cell.value).lower()
            if var == 'none':
                if str(wb.cell(counter,16).value) == 'None':
                    continue
                else:
                    unknown.append(var)
                    continue
            elif var == 'unknown':
                unknown.append(var)
                continue
            elif var == '2/1/1857':
                years_start.append('1857')
                years_end.append('2022')
            elif str(wb.cell(counter,15).value) == '00:00:00':
                continue
            else:
                if str(wb.cell(counter,16).value) == 'None' or str(wb.cell(counter,16).value).lower() == 'unknown':
                    unknown.append(var)
                    continue
                date_s = datetime.strptime(var,date_format)
                years_start.append(date_s.year)
                date_e = datetime.strptime(str(wb.cell(counter,16).value),date_format)
                years_end.append(date_e.year)
                continue
except:             #handles any exceptions outside of parameters
    print(counter)
    raise

#------------------------------------------------------
#checking for streamtype values outside of standard
#------------------------------------------------------
for row in wb.iter_rows(min_col=9,max_col=9,min_row=2,max_row=maxrow):
    for cell in row:
        counter += 1
        var = str(cell.value).lower()
        var = var.rstrip(' ')
        if str(wb.cell(counter,21).value) == 'continuous': #subset by continuous
            if str(wb.cell(counter,1).value) == 'None':
                continue
            elif var == 'none' or var == 'unknown':
                if wb.cell(counter,6).value == I:
                    ID_unk.append(var)
                elif wb.cell(counter,6).value == O:
                    OR_unk.append(var)
                elif wb.cell(counter,6).value == W:
                    WA_unk.append(var)
            elif var == 'canal':
                if wb.cell(counter,6).value == I:
                    ID_canal.append(var)
                elif wb.cell(counter,6).value == O:
                    OR_canal.append(var)
                elif wb.cell(counter,6).value == W:
                    WA_canal.append(var)
            elif var == 'stream':
                if wb.cell(counter,6).value == I:
                    ID_stream.append(var)
                elif wb.cell(counter,6).value == O:
                    OR_stream.append(var)
                elif wb.cell(counter,6).value == W:
                    WA_stream.append(var)
            elif var == 'weir':
                if wb.cell(counter,6).value == I:
                    ID_weir.append(var)
                elif wb.cell(counter,6).value == O:
                    OR_stream.append(var)
                elif wb.cell(counter,6).value == W:
                    WA_stream.append(var)
        else:
            exceptions.append(str(counter)+ var)
